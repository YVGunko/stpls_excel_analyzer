import pandas as pd
import os
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

# Define a thick top border style
thick_top_border = Border(top=Side(style="thick"))
first_table_row = 6

def trim_product_name(product_name):
    if not isinstance(product_name, str) or not product_name.strip():
        return ""
    
    words = product_name.split()
    
    if len(words) < 3:
        return product_name  # No need to adjust if there aren't enough words

    if words[0].upper().startswith("ПОДОШВА") and len(words) > 3:
        # Check if the second word is not in the list, but the third one is
        key_words = {"ПУ", "ТЭП", "ЭВА"}
        if words[1].upper() not in key_words and words[2].upper() in key_words:
            words[1], words[2] = words[2], words[1]  # Swap second and third words
        return " ".join(word.capitalize() for word in words[:3])
    elif words[0].upper().startswith("СТЕЛЬКА") and len(words) > 2:
        return " ".join(word.capitalize() for word in words[:2])
    
    return " ".join([word.capitalize() for word in words])

def convert_xls_to_xlsx(input_path):
    if input_path.endswith(".xls"):
        output_path = input_path + "x"  # Change ".xls" to ".xlsx"
        
        # Load old .xls file
        xls = pd.ExcelFile(input_path)
        
        # Save as new .xlsx file
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"Converted {input_path} → {output_path}")
        return output_path  # Return new .xlsx path
    return input_path  # If already .xlsx, return as is

def analyze_excel():
    # Open file dialog to select input Excel file
    Tk().withdraw()  # Hide the root window
    input_path = filedialog.askopenfilename(title="Выберите файл Excel", filetypes=[("Excel Files", "*.xlsx *.xls")])
    input_path = convert_xls_to_xlsx(input_path)  # Ensure it's .xlsx
    if not input_path:
        print("No file selected. Exiting.")
        return
    
    # Generate output file name
    output_path = os.path.join(os.path.dirname(input_path), "s_" + os.path.basename(input_path))
    
    # Load the Excel file
    xls = pd.ExcelFile(input_path)
    df = pd.read_excel(xls, sheet_name=xls.sheet_names[0], dtype=str)
    
    header_texts = {"Накладная": "", "Поставщик:": "", "Покупатель:": ""}
    for row in df.iloc[:7].values:  
        for i, cell in enumerate(row):
            if isinstance(cell, str):
                for key in header_texts:
                    if key in cell:
                        if key == "Накладная":
                            header_texts[key] = cell.strip()
                        else:
                            value = ""
                            for j in range(i + 1, len(row)):  
                                if isinstance(row[j], str) and row[j].strip():
                                    value = row[j].strip()
                                    break  
                            header_texts[key] = f"{cell.strip()} {value}".strip()

    # **Find footer texts (from last 7 rows)**
    footer_texts = {"Итого:": "", "НДС:": ""}
    for row in df.iloc[-8:].values:  # Scan last 7 rows
        for i, cell in enumerate(row):
            if isinstance(cell, str):
                for key in footer_texts:
                    if key in cell:
                        value = ""
                        for j in range(i + 1, len(row)):  
                            if isinstance(row[j], str) and row[j].strip():
                                value = row[j].strip()
                                break  
                        footer_texts[key] = f"{cell.strip()} {value}".strip()    
    # Identify the correct header row
    header_row_index = 7
    df_headers = pd.read_excel(xls, sheet_name=xls.sheet_names[0], skiprows=header_row_index, nrows=1, dtype=str)
    clean_headers = df_headers.iloc[0].fillna("").astype(str).tolist()
    df = pd.read_excel(xls, sheet_name=xls.sheet_names[0], skiprows=header_row_index + 1, dtype=str)
    df.columns = clean_headers
    
    # Find the first row where analysis should start
    first_data_row = df[df.iloc[:, 1] == "1"].index.min()
    
    # Extract required columns
    order_col = df.columns[df.columns.str.contains("№", case=False, na=False)][0]
    product_col = df.columns[df.columns.str.contains("Товар", case=False, na=False)][0]
    quantity_col = df.columns[df.columns.str.contains("Количество", case=False, na=False)][0]
    places_col = df.columns[df.columns.str.contains("Мест", case=False, na=False)][0]
    total_col = df.columns[df.columns.str.contains("Сумма", case=False, na=False)][0]
    
    # Process relevant rows
    df_data = df.iloc[first_data_row:].reset_index(drop=True)
    # Apply trimming and reordering for proper grouping
    df_data[product_col] = df_data[product_col].apply(lambda name: trim_product_name(name))

    # Sort rows by the cleaned "Товар" column (case-insensitive)
    df_data = df_data.sort_values(by=product_col, key=lambda col: col.str.lower() if col.dtype == "object" else col)

    # Group similar product names
    results = []
    current_order = 1
    current_product = ""
    current_places = "пар."
    total_quantity = 0
    total_sum = 0
    print (results)
    for _, row in df_data.iterrows():
        product_name = row[product_col]

        if len (product_name.strip()) == 0 : continue
        quantity = row[quantity_col].split()[0] if pd.notna(row[quantity_col]) else "0"
        total_value = row[total_col] if pd.notna(row[total_col]) else "0"
        
        if product_name != current_product and current_product:
            results.append([current_order, current_product, current_places, total_quantity, total_sum])
            print (results)
            current_order = current_order + 1
            total_quantity, total_sum = 0, 0
        
        current_product = product_name
        total_quantity += int(quantity)
        total_sum += float(total_value)
    
    if current_product:
        results.append([current_order, current_product, current_places, total_quantity, total_sum])
    
    # Create output DataFrame (WITHOUT inserting extra rows yet)
    output_df = pd.DataFrame(results, columns=["№", "Товар", "Ед.изм.", "Количество", "Сумма"])
    print (output_df)
    output_df.to_excel(output_path, index=False, engine="openpyxl")

    # Now modify the file with openpyxl
    wb = load_workbook(output_path)
    ws = wb.active

    # **Insert header texts (without shifting the analysis)**
    ws.insert_rows(1, amount=4)
    ws["B1"], ws["B2"], ws["B3"] = header_texts["Накладная"], header_texts["Поставщик:"], header_texts["Покупатель:"]

    # Determine the last row with data
    last_row = ws.max_row
    # Apply thick top border to all columns in the summary row
    for col in ["B", "C", "D", "E"]:  # Columns from "№" to "Сумма"
        ws[f"{col}{last_row + 1}"].border = thick_top_border
    # Insert SUM formula in the next row for columns 4 and 5
    ws[f"B{last_row + 1}"] = "Итого:"  # Итого
    ws[f"D{last_row + 1}"] = f"=SUM(D{first_table_row}:D{last_row})"  # Количество
    ws[f"E{last_row + 1}"] = f"=SUM(E{first_table_row}:E{last_row})"  # Сумма

    # Apply number formatting to match the table
    ws[f"D{last_row + 1}"].number_format = "#,##0.00"
    ws[f"E{last_row + 1}"].number_format = "#,##0.00"

    # **Insert footer texts into column B**
    footer_row = ws.max_row + 2  # Leave an empty row before footer
    ws[f"B{footer_row}"] = footer_texts["Итого:"]
    ws[f"B{footer_row + 1}"] = footer_texts["НДС:"]

    # **Apply bold styling to headers (corrected row number)**
    for cell in ws[4]:  # Headers now on row 4
        cell.font = Font(bold=True)

    # Apply number format to "Количество" and "Сумма" columns
    for row in ws.iter_rows(min_row=first_table_row, max_row=ws.max_row, min_col=4, max_col=5):  # Columns D & E
        for cell in row:
            cell.number_format = "#,##0.00"

    # **Auto-adjust column width**
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2



    # Save final result
    wb.save(output_path)

    print(f"Analysis complete. Output saved to {output_path}")


# Run the function
analyze_excel()
